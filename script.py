#from tabnanny import filename_only
from tkinter import messagebox
from tkinter.simpledialog import askstring
from openpyxl import Workbook, load_workbook
from srtools import *
from tkinter import filedialog as fd
import tkinter as tk
from tkinter.messagebox import showinfo
import os

top = tk.Tk()
top.title("Program za ažuriranje cena") 
top.geometry("500x200")

def sredi_cene():
    promena_opisa()
    poredjenje_cena()
    promena_cene_zalihe()
    promena_cene_stanje()
    save_excel()

def ucitaj():
    global fisherman_front, zalihe_front, kategorije_front
    global fisherman_book, zalihe_book, kategorije_book
    global fisherman, zalihe, kategorije
    global naslov_fisherman, cena_fisherman, kategorija_fisherman, opis_fisherman, sifra_fisherman
    global artikal_zalihe, naziv_zalihe, cena_zalihe, stanje_zalihe
    global artikal_kategorije, kategorija_kategorije
    global artikal_list, sifra_list, stanje_list

    # ucitavanje excel fajlova na osnovu izabranih fajlova u front.py 
    fisherman_book = load_workbook(filename=fisherman_name)
    zalihe_book = load_workbook(filename=zalihe_name)
    kategorije_book = load_workbook(filename=kategorije_name)
    fisherman = fisherman_book.active
    zalihe = zalihe_book.active
    kategorije = kategorije_book.active
    # access the active worksheet data
    naslov_fisherman = fisherman["A"]
    sifra_fisherman = fisherman["B"]
    cena_fisherman = fisherman["C"]
    kategorija_fisherman = fisherman["K"]
    opis_fisherman = fisherman["H"]
    cena_zalihe = zalihe["D"]
    artikal_zalihe = zalihe["A"]
    stanje_zalihe = zalihe["C"]
    naziv_zalihe = zalihe["B"]
    artikal_kategorije = kategorije["A"]
    kategorija_kategorije = kategorije["D"]
    # store cell values in a list for comparing
    sifra_list = []
    artikal_list = []
    stanje_list = []

    for row in sifra_fisherman:
        sifra_list.append(row.value.strip())

    for row in artikal_zalihe:
        artikal_list.append(row.value.strip())

    for row in stanje_zalihe:
        stanje_list.append(row.value)

# uvoz izabranih fajlova
filetypes = (('excel', '*.xls'), ('excel', '*.xlsx'), ('excel', '*.xlsm'))

def uvoz_fisherman():
    global fisherman_front, fisherman_name
    filename = fd.askopenfilename(title='Open a file', initialdir='/Desktop', filetypes=filetypes)
    fisherman_front = filename
    fisherman_name = os.path.basename(fisherman_front)
    print(fisherman_name)

def uvoz_zalihe():
    global zalihe_front, zalihe_name
    filename = fd.askopenfilename(title='Open a file', initialdir='/Desktop', filetypes=filetypes)
    zalihe_front = filename
    zalihe_name = os.path.basename(zalihe_front)
    print(zalihe_name)

def uvoz_kategorije():
    global kategorije_front, kategorije_name
    filename = fd.askopenfilename(title='Open a file', initialdir='/Desktop', filetypes=filetypes)
    kategorije_front = filename
    kategorije_name = os.path.basename(kategorije_front)
    print(kategorije_name)

# provera kolone opis i promena cirilice u latinicu
def promena_opisa():
    for i in range(opis_fisherman.__len__()):
        if opis_fisherman[i].value != None:
            opis_fisherman[i].value = cyrillic_to_latin(opis_fisherman[i].value)
            #print("stara vrednost: " + opis_fisherman[i].value + " nova vrednost: " + cyrillic_to_latin(opis_fisherman[i].value))
        else:
            continue

# novi proizvodi koji se nalaze u zalihama ali nisu u fishermanu 
def novi_proizvodi():
    global artikal_list, sifra_list, novi_proizvodi_list, stanje_list
    global artikal_zalihe, naziv_zalihe, cena_zalihe, stanje_zalihe
    global fisherman
    novi_proizvodi_list = []
    for i in range(1,artikal_list.__len__()):
        if artikal_list[i] not in sifra_list:
            print("artikal: " + artikal_zalihe[i].value + "\nnaziv: " + naziv_zalihe[i].value + "\nstanje: " + str(stanje_zalihe[i].value) + "\ncena: " + str(cena_zalihe[i].value))
            novi_proizvodi_list.append([artikal_zalihe[i].value, naziv_zalihe[i].value, stanje_zalihe[i].value, cena_zalihe[i].value])
            #novi proizvod moze da ima stanje manje od 0 ako ovu funkciju preklopis sa funkcijom promena_cene_stanje()
            #if stanje_zalihe[i].value > 0:
            fisherman.cell(row=fisherman.max_row + 1, column=1).value = naziv_zalihe[i].value
            fisherman.cell(row=fisherman.max_row , column=2).value = artikal_zalihe[i].value
            fisherman.cell(row=fisherman.max_row , column=3).value = cena_zalihe[i].value
            fisherman.cell(row=fisherman.max_row , column=4).value = "kom"
            fisherman.cell(row=fisherman.max_row , column=13).value = "2021-04-30 00:00:13"

    if novi_proizvodi_list.__len__() != 0:                
        print("lista novih proizvoda: " + str(novi_proizvodi_list))
        #izrada izvestaja o novim proizvodima
        report = messagebox.askquestion("Report", "Da li zelite da sacuvate report u excel fajl?")
        if report == 'yes':
            report_book = Workbook()
            report_sheet = report_book.active
            report_sheet.append(["artikal", "naziv", "stanje", "cena", "kategorija"])
            for i in range(novi_proizvodi_list.__len__()):
                report_sheet.append(novi_proizvodi_list[i])
            report_book.save("report.xlsx")
            showinfo(title='Report', message="Report je sacuvan u report.xlsx fajlu")
    else:
        showinfo(title='Report', message="Nema novih proizvoda")
    save_excel()

# funkcija za proveru i promenu kategorije
def promena_kategorije():
    for i in range(artikal_kategorije.__len__()):
        for j in range(sifra_fisherman.__len__()):
            if artikal_kategorije[i].value == sifra_fisherman[j].value and kategorija_fisherman[j].value != kategorija_kategorije[i].value:
                print('sifra proizvoda: ' + str(sifra_fisherman[j].value) + ' stara kategorija: ' + str(kategorija_fisherman[j].value) + ' nova kategorija: ' + str(kategorija_kategorije[i].value))
                kategorija_fisherman[j].value = kategorija_kategorije[i].value
            else:
                continue
    save_excel()

# proizvodi kojih nema na zalihama ili su na zalihama ali ih nema na stanju, promeni cenu na nulu
# funkcije su razdvojene zbog performansi
def promena_cene_zalihe():
    for i in range(sifra_fisherman.__len__()):
        if sifra_fisherman[i].value not in artikal_list and cena_fisherman[i].value != 0:
            print("fisherman: " + sifra_fisherman[i].value + " cena: " + str(cena_fisherman[i].value) + " nova cena zalihe: 0")
            cena_fisherman[i].value = 0
        else:
            continue
    save_excel()

def promena_cene_stanje():
    for i in range(sifra_fisherman.__len__()):
        for j in range(artikal_zalihe.__len__()):
            if sifra_fisherman[i].value == artikal_zalihe[j].value and cena_fisherman[i].value != 0 and stanje_zalihe[j].value < 1:
                print("fisherman: " + sifra_fisherman[i].value + " cena: " + str(cena_fisherman[i].value) + " nova cena stanje: 0")
                cena_fisherman[i].value = 0
            else:
                continue
    save_excel()

# funkcija za poredjenje cena na zalihama i u fishermanu
def poredjenje_cena():
    for i in range(sifra_fisherman.__len__()):
        for j in range(artikal_zalihe.__len__()):
            if sifra_fisherman[i].value == artikal_zalihe[j].value and cena_fisherman[i].value != cena_zalihe[j].value:
                print("fisherman: " + sifra_fisherman[i].value + " cena: " + str(cena_fisherman[i].value) + " nova cena: " + str(cena_zalihe[j].value))
                cena_fisherman[i].value = cena_zalihe[j].value
    save_excel()

# pretraga proizvoda po sifri
def pretraga_po_sifri():
    value = askstring("Pretraga", "Unesite trazeni pojam: ")
    flag = 1
    for i in range(sifra_fisherman.__len__()):
        # if value in sifra_fisherman[i].value or value in naslov_fisherman[i].value: # ako sadrzi trazeni pojam
        if value == sifra_fisherman[i].value or value == naslov_fisherman[i].value:   # ako je tacan trazeni pojam
            print("naziv: " + naslov_fisherman[i].value + "\nsifra: " + sifra_fisherman[i].value + "\ncena: " + str(cena_fisherman[i].value))
            showinfo(title='Pretraga', message="naziv: " + naslov_fisherman[i].value + "\nsifra: " + sifra_fisherman[i].value + "\ncena: " + str(cena_fisherman[i].value))
            flag = 1
            break
        elif value == '':
            showinfo(title='Pretraga', message="Niste uneli pojam za pretragu")
            break
        else:
            #showinfo(title='Pretraga', message="Nema rezultata")
            flag = 0
            continue
    if flag == 0:
        showinfo(title='Pretraga', message="Nema rezultata")

# zapamti u excel
def save_excel():
    fisherman_book.save(fisherman_name)
    zalihe_book.save(zalihe_name)
    kategorije_book.save(kategorije_name)

btn_sredi_cene = tk.Button(top, text="Sredi cene", command=sredi_cene)
btn_sredi_cene.grid(row=12, column=0, padx=10, pady=10)
btn_novi_proizvodi = tk.Button(top, text="Novi proizvodi", command=novi_proizvodi)
btn_novi_proizvodi.grid(row=12, column=2, padx=10, pady=10)
btn_pretraga = tk.Button(top, text="Pretraži" , command=pretraga_po_sifri)
btn_pretraga.grid(row=12, column=3, padx=10, pady=10)
btn_import_fisherman = tk.Button(top, text="Import Fisherman", command=uvoz_fisherman)
btn_import_fisherman.grid(row=3, column=0, padx=20, pady=30)
btn_import_zalihe = tk.Button(top, text="Import Zalihe", command=uvoz_zalihe)
btn_import_zalihe.grid(row=3, column=1, padx=10, pady=10)
btn_import_kategorije = tk.Button(top, text="Import Kategorije", command=uvoz_kategorije)
btn_import_kategorije.grid(row=3, column=2, padx=10, pady=10)
btn_ucitaj = tk.Button(top, text="Učitaj podatke", command=ucitaj)
btn_ucitaj.grid(row=3, column=3, padx=10, pady=10)
btn_sredi_kategorije = tk.Button(top, text="Sredi kategorije", command=promena_kategorije)
btn_sredi_kategorije.grid(row=12, column=1, padx=10, pady=10)
top.mainloop()